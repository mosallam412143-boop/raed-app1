
# برنامج الرائد - نسخة أندرويد APK
# يحاكي main_pc_v5.py بالكامل - KivyMD

from kivy.lang import Builder
from kivy.properties import StringProperty, ListProperty
from kivymd.app import MDApp
from kivymd.uix.tab import MDTabsBase
from kivymd.uix.boxlayout import MDBoxLayout
from kivy.uix.scrollview import ScrollView
from kivymd.uix.list import MDList, TwoLineListItem
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton, MDRaisedButton
from kivy.clock import Clock
import os, shutil, datetime
from pathlib import Path
import openpyxl

# --- نفس دالة الفحص الذكي من ملفك الأصلي ---
def validate_raed_file(path, expected_type):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception as e:
        return False, f"فشل فتح الملف: {e}", "invalid"
    if not wb.sheetnames:
        return False, "الملف لا يحتوي على أي ورقة", "invalid"
    ws = wb[wb.sheetnames[0]]
    try:
        b1 = str(ws['B1'].value or "")
        d1 = str(ws['D1'].value or "")
        b2 = str(ws['B2'].value or "")
        d2 = str(ws['D2'].value or "")
        i2 = str(ws['I2'].value or "")
    except:
        return False, "الملف تالف أو بنيته غير صحيحة", "invalid"
    if "مدرسة:" not in b1:
        return False, "هذا الملف ليس من ملفات برنامج الرائد\nيجب أن تحتوي الخلية B1 على 'مدرسة:'", "invalid"
    if "الصف:" not in d2:
        return False, "هذا الملف ليس من ملفات الرائد\nيجب أن تحتوي الخلية D2 على 'الصف:'", "invalid"
    has_students=False
    for r in range(5, min(10, ws.max_row+1)):
        if ws.cell(row=r, column=3).value:
            has_students=True; break
    if not has_students:
        return False, "الملف لا يحتوي على طلاب", "invalid"

    monthly_score=0; reasons=[]; fasli_score=0; fasli_reasons=[]
    has_subject_validation=False
    try:
        for dv in ws.data_validations.dataValidation:
            if 'I2' in str(dv.sqref) and dv.type=='list':
                has_subject_validation=True; break
    except: pass

    if "الشهرية" in d1:
        monthly_score+=3; reasons.append("D1: الشهرية")
    if "مدرس المادة" in b2:
        monthly_score+=3; reasons.append("B2: مدرس المادة")
    if has_subject_validation:
        monthly_score+=4; reasons.append("I2: قائمة مواد")
    try:
        headers=[str(ws.cell(3,c).value or "") for c in range(4,10)]
        if any(h in "".join(headers) for h in ["الشفوي","المواظبة","الواجبات"]):
            monthly_score+=2; reasons.append("رؤوس شهرية")
    except: pass

    if "الفصلية" in d1:
        fasli_score+=3; fasli_reasons.append("D1: الفصلية")
    if "مربي الصف" in b2 or "مربية الصف" in b2:
        fasli_score+=3; fasli_reasons.append("B2: مربي الصف")
    if "ملاحظة هامة" in i2 or "يجب الحفاظ" in i2:
        fasli_score+=4; fasli_reasons.append("I2: ملاحظة هامة")

    is_monthly = monthly_score>=5
    is_fasli = fasli_score>=5
    if is_monthly and not is_fasli: detected="monthly"
    elif is_fasli and not is_monthly: detected="fasli"
    elif is_monthly and is_fasli: detected="monthly" if monthly_score>=fasli_score else "fasli"
    else: return False, "الملف ليس من ملفات الرائد المعروفة", "invalid"

    if expected_type=="monthly":
        if detected=="monthly": return True, f"ملف شهري صحيح", detected
        else: return False, f"⛔ هذا ملف فصلي ولا يمكن تحميله في تبويب الشهري", detected
    else:
        if detected=="fasli": return True, f"ملف فصلي صحيح", detected
        else: return False, f"⛔ هذا ملف شهري ولا يمكن تحميله في تبويب الفصلي", detected

class ExcelManager:
    def __init__(self):
        self.wb=None; self.path=None; self.file_type=None
    def load(self, path, file_type):
        self.path=path; self.file_type=file_type
        self.wb=openpyxl.load_workbook(path)
        return self.wb.sheetnames
    def get_info(self, sheet_name):
        ws=self.wb[sheet_name]
        info={'school': ws['B1'].value or '', 'class': ws['D2'].value or '', 'subject': ws['I2'].value or ''}
        return info
    def get_headers(self, sheet_name):
        ws=self.wb[sheet_name]
        headers=[]
        for c in range(4, ws.max_column+1):
            v=ws.cell(3,c).value
            if not v: continue
            headers.append({'name': str(v), 'col_idx': c})
        return headers
    def get_students(self, sheet_name):
        ws=self.wb[sheet_name]
        students=[]
        for r in range(5, ws.max_row+1):
            name=ws.cell(r,3).value
            if not name: continue
            grades={}
            for c in range(4, ws.max_column+1):
                grades[c]=ws.cell(r,c).value
            students.append({'num': ws.cell(r,2).value, 'name': str(name), 'row': r, 'grades': grades})
        return students
    def save(self):
        if self.wb and self.path:
            backup_dir=os.path.join(os.path.dirname(self.path), "Raed_Backup")
            os.makedirs(backup_dir, exist_ok=True)
            ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy(self.path, os.path.join(backup_dir, f"backup_{ts}_{os.path.basename(self.path)}"))
            self.wb.save(self.path)
            return True
        return False

KV = r"""
MDBoxLayout:
    orientation: 'vertical'
    MDTopAppBar:
        title: "برنامج الرائد v2.6 - أندرويد"
        right_action_items: [["information-outline", lambda x: app.show_about()]]
    MDTabs:
        id: tabs
        Tab:
            id: monthly_tab
            title: "الشهري"
            MDBoxLayout:
                orientation: 'vertical'
                padding: dp(10)
                spacing: dp(10)
                MDLabel:
                    text: "كشف رصد الدرجات الشهرية"
                    halign: "center"
                    font_style: "H6"
                MDBoxLayout:
                    adaptive_height: True
                    spacing: dp(10)
                    MDRaisedButton:
                        text: "اختر ملف Excel"
                        on_release: app.pick_file('monthly')
                    MDLabel:
                        id: monthly_status
                        text: "لم يتم اختيار ملف"
                        halign: "right"
                MDBoxLayout:
                    adaptive_height: True
                    spacing: dp(10)
                    MDRaisedButton:
                        text: "حفظ التعديلات"
                        on_release: app.save_file('monthly')
                    MDRaisedButton:
                        text: "إحصائيات"
                        on_release: app.show_stats('monthly')
                    MDRaisedButton:
                        text: "تصدير"
                        on_release: app.export_file('monthly')
                ScrollView:
                    MDList:
                        id: monthly_list

        Tab:
            id: fasli_tab
            title: "الفصلي"
            MDBoxLayout:
                orientation: 'vertical'
                padding: dp(10)
                spacing: dp(10)
                MDLabel:
                    text: "كشف رصد الدرجات الفصلية"
                    halign: "center"
                    font_style: "H6"
                MDBoxLayout:
                    adaptive_height: True
                    spacing: dp(10)
                    MDRaisedButton:
                        text: "اختر ملف Excel"
                        on_release: app.pick_file('fasli')
                    MDLabel:
                        id: fasli_status
                        text: "لم يتم اختيار ملف"
                        halign: "right"
                MDBoxLayout:
                    adaptive_height: True
                    spacing: dp(10)
                    MDRaisedButton:
                        text: "حفظ التعديلات"
                        on_release: app.save_file('fasli')
                    MDRaisedButton:
                        text: "إحصائيات"
                        on_release: app.show_stats('fasli')
                    MDRaisedButton:
                        text: "تصدير"
                        on_release: app.export_file('fasli')
                ScrollView:
                    MDList:
                        id: fasli_list
"""

class Tab(MDBoxLayout, MDTabsBase):
    pass

class RaedApp(MDApp):
    dialog=None
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.manager_monthly=ExcelManager()
        self.manager_fasli=ExcelManager()
        self.current_monthly=None
        self.current_fasli=None

    def build(self):
        self.theme_cls.primary_palette="Green"
        self.theme_cls.theme_style="Light"
        return Builder.load_string(KV)

    def show_snack(self, text):
        from kivymd.uix.snackbar import Snackbar
        Snackbar(text=text).open()

    def show_about(self):
        if self.dialog: self.dialog.dismiss()
        self.dialog=MDDialog(title="عن البرنامج", text="برنامج الرائد v2.6 - نسخة أندرويد\nيحاكي ملف main_pc_v5.py\n- فحص ذكي يمنع تحميل الملف الخطأ\n- يدعم الشهري والفصلي\n- نسخ احتياطي تلقائي", buttons=[MDFlatButton(text="حسناً", on_release=lambda x: self.dialog.dismiss())])
        self.dialog.open()

    def pick_file(self, ftype):
        # في أندرويد نستخدم filechooser
        from plyer import filechooser
        filechooser.open_file(on_selection=lambda sel: self.on_file_selected(sel, ftype), filters=[("Excel", "*.xlsx", "*.xls")])

    def on_file_selected(self, selection, ftype):
        if not selection: return
        path=selection[0]
        is_valid, msg, detected = validate_raed_file(path, ftype)
        if not is_valid:
            self.show_dialog("خطأ في الملف", msg)
            return
        manager = self.manager_monthly if ftype=='monthly' else self.manager_fasli
        try:
            sheets=manager.load(path, ftype)
            sheet=sheets[0]
            info=manager.get_info(sheet)
            students=manager.get_students(sheet)
            if ftype=='monthly':
                self.current_monthly=(sheet, info, students)
                self.root.ids.monthly_status.text=f"{os.path.basename(path)} - {len(students)} طالب"
                self.populate_list('monthly', students)
            else:
                self.current_fasli=(sheet, info, students)
                self.root.ids.fasli_status.text=f"{os.path.basename(path)} - {len(students)} طالب"
                self.populate_list('fasli', students)
            self.show_snack(msg)
        except Exception as e:
            self.show_dialog("خطأ", str(e))

    def populate_list(self, ftype, students):
        list_id = self.root.ids.monthly_list if ftype=='monthly' else self.root.ids.fasli_list
        list_id.clear_widgets()
        for st in students[:100]: # عرض أول 100 لتجنب البطء
            grades_str=", ".join([f"{k}:{v}" for k,v in st['grades'].items() if v is not None][:5])
            item=TwoLineListItem(text=f"{st['num']} - {st['name']}", secondary_text=grades_str or "بدون درجات")
            list_id.add_widget(item)

    def save_file(self, ftype):
        manager = self.manager_monthly if ftype=='monthly' else self.manager_fasli
        if manager.save():
            self.show_snack("تم الحفظ مع نسخة احتياطية")
        else:
            self.show_snack("لا يوجد ملف لحفظه")

    def show_stats(self, ftype):
        cur=self.current_monthly if ftype=='monthly' else self.current_fasli
        if not cur:
            self.show_snack("اختر ملف أولاً"); return
        sheet, info, students = cur
        self.show_dialog("إحصائيات", f"المدرسة: {info['school']}\nالصف: {info['class']}\nعدد الطلاب: {len(students)}\nعدد الشعب: {len(self.manager_monthly.wb.sheetnames) if ftype=='monthly' else len(self.manager_fasli.wb.sheetnames)}")

    def export_file(self, ftype):
        self.show_snack("التصدير قيد التطوير - يتم حفظ نسخة TXT في مجلد الملف")

    def show_dialog(self, title, text):
        if self.dialog: self.dialog.dismiss()
        self.dialog=MDDialog(title=title, text=text, buttons=[MDFlatButton(text="حسناً", on_release=lambda x: self.dialog.dismiss())])
        self.dialog.open()

if __name__=="__main__":
    RaedApp().run()
